import os
import re
import json
import asyncio
from slack_bolt.async_app import AsyncApp
from slack_bolt.adapter.socket_mode.async_handler import AsyncSocketModeHandler
import google.generativeai as genai
from dotenv import load_dotenv

# Tải cấu hình từ file .env
load_dotenv()

# Cấu hình API Key của Google Gemini
genai.configure(api_key=os.environ.get("GEMINI_API_KEY"))
import requests

# Đường dẫn thư mục tài liệu SRS
SRS_DIR = os.environ.get("SRS_DIR", "/home/phamhung/Work/MVL/web/docs/srs/docs")

def call_clickup_mcp(tool_name, arguments):
    import subprocess
    token = os.environ.get("CLICKUP_TOKEN", "pk_288804163_TS4QZV0GZEDMO5MNYVMY03FR1QSPLJNS")
    cmd = ["npx", "-y", "@cavort-it-systems/clickup-mcp"]
    env = dict(os.environ, CLICKUP_API_TOKEN=token)
    proc = subprocess.Popen(
        cmd,
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        bufsize=1,
        env=env
    )
    try:
        # 1. Initialize
        init_req = {
            "jsonrpc": "2.0",
            "id": 1,
            "method": "initialize",
            "params": {
                "protocolVersion": "2024-11-05",
                "capabilities": {},
                "clientInfo": {"name": "workflow-client", "version": "1.0"}
            }
        }
        proc.stdin.write(json.dumps(init_req) + "\n")
        proc.stdin.flush()
        
        init_resp = None
        for line in proc.stdout:
            try:
                data = json.loads(line)
                if data.get("id") == 1:
                    init_resp = data
                    break
            except json.JSONDecodeError:
                continue
                
        if not init_resp:
            return {"error": "Failed to initialize MCP server"}
            
        # 2. Send initialized notification
        init_notif = {
            "jsonrpc": "2.0",
            "method": "notifications/initialized"
        }
        proc.stdin.write(json.dumps(init_notif) + "\n")
        proc.stdin.flush()
        
        # 3. Call tool
        tool_req = {
            "jsonrpc": "2.0",
            "id": 2,
            "method": "tools/call",
            "params": {
                "name": tool_name,
                "arguments": arguments
            }
        }
        proc.stdin.write(json.dumps(tool_req) + "\n")
        proc.stdin.flush()
        
        # 4. Read response
        tool_resp = None
        for line in proc.stdout:
            try:
                data = json.loads(line)
                if data.get("id") == 2:
                    tool_resp = data
                    break
            except json.JSONDecodeError:
                continue
                
        if tool_resp and "result" in tool_resp:
            content = tool_resp["result"].get("content", [])
            if content and content[0].get("type") == "text":
                try:
                    return json.loads(content[0]["text"])
                except Exception:
                    return content[0]["text"]
        return tool_resp
    except Exception as e:
        return {"error": str(e)}
    finally:
        proc.stdin.close()
        proc.terminate()
        proc.wait()

def search_clickup_tasks(query: str, filter_bugs: bool = None) -> str:
    """
    Tìm kiếm các công việc (tasks) hoặc bug trên ClickUp liên quan đến từ khóa sử dụng ClickUp MCP.

    Args:
        query: Từ khóa hoặc chủ đề cần tìm kiếm (ví dụ: 'tạm ứng', 'hoa hồng', 'đối chiếu').
        filter_bugs: Đặt là True nếu chỉ muốn tìm bug/lỗi. Đặt là False nếu chỉ muốn tìm task nghiệp vụ. Đặt là None/không truyền để lấy cả hai.
    """
    team_id = os.environ.get("CLICKUP_TEAM_ID", "90181237095")
    res = call_clickup_mcp("clickup_search_tasks", {
        "team_id": team_id,
        "search": query,
        "subtasks": True,
        "include_closed": True
    })
    
    if isinstance(res, dict) and "error" in res:
        return f"Lỗi khi tìm kiếm task trên ClickUp bằng MCP: {res['error']}"
        
    tasks = []
    if isinstance(res, dict):
        tasks = res.get("tasks", [])
    elif isinstance(res, list):
        tasks = res
        
    if not tasks:
        return f"Không tìm thấy công việc nào trên ClickUp."
        
    def remove_accents(input_str: str) -> str:
        import unicodedata
        input_str = input_str.replace('đ', 'd').replace('Đ', 'D')
        nfkd_form = unicodedata.normalize('NFKD', input_str)
        return "".join([c for c in nfkd_form if not unicodedata.combining(c)])
        
    # Lọc theo từ khóa tìm kiếm (query) ở phía Python vì MCP không hỗ trợ tham số search
    if query:
        search_keywords = [k.strip().lower() for k in query.split() if len(k.strip()) >= 2]
        if search_keywords:
            filtered_tasks = []
            for t in tasks:
                name = t.get("name", "").lower()
                desc = (t.get("description") or "").lower()
                
                name_no_accents = remove_accents(name)
                desc_no_accents = remove_accents(desc)
                
                match = True
                for kw in search_keywords:
                    kw_no_accents = remove_accents(kw)
                    kw_pattern = rf'\b{re.escape(kw)}\b'
                    kw_no_accents_pattern = rf'\b{re.escape(kw_no_accents)}\b'
                    
                    in_accented = bool(re.search(kw_pattern, name)) or bool(re.search(kw_pattern, desc))
                    in_unaccented = bool(re.search(kw_no_accents_pattern, name_no_accents)) or bool(re.search(kw_no_accents_pattern, desc_no_accents))
                    
                    if not (in_accented or in_unaccented):
                        match = False
                        break
                if match:
                    filtered_tasks.append(t)
            tasks = filtered_tasks

    # Lọc theo loại công việc (bug vs task)
    if filter_bugs is not None:
        filtered_tasks = []
        for t in tasks:
            name = t.get("name", "").lower()
            desc = (t.get("description") or "").lower()
            is_bug = "bug" in name or "bug" in desc or "lỗi" in name or "lỗi" in desc
            if filter_bugs == is_bug:
                filtered_tasks.append(t)
        tasks = filtered_tasks
        
    if not tasks:
        return f"Không tìm thấy kết quả phù hợp sau khi lọc theo yêu cầu."
        
    result = []
    title = "Các bug/lỗi liên quan trên ClickUp:" if filter_bugs else "Các công việc liên quan trên ClickUp:"
    result.append(f"🔍 {title}")
    for t in tasks[:10]:
        t_id = t.get("id")
        t_name = t.get("name")
        t_status = t.get("status", {}).get("status", "N/A")
        t_url = t.get("url")
        result.append(f"- [{t_id}] {t_name} - Trạng thái: {t_status}")
        
    return "\n".join(result)


def fetch_clickup_task_direct(task_id: str) -> dict:
    token = os.environ.get("CLICKUP_TOKEN", "pk_288804163_TS4QZV0GZEDMO5MNYVMY03FR1QSPLJNS")
    headers = {"Authorization": token}
    url = f"https://api.clickup.com/api/v2/task/{task_id}"
    try:
        resp = requests.get(url, headers=headers, timeout=5)
        if resp.status_code == 200:
            return resp.json()
    except Exception:
        pass
    return None


def get_clickup_task(task_id: str) -> str:
    """
    Lấy thông tin chi tiết của một task hoặc bug cụ thể trên ClickUp theo ID của task.
    
    Args:
        task_id: ID hoặc URL của task ClickUp (ví dụ: '86exxz2xr').
    """
    import concurrent.futures
    
    # Làm sạch và trích xuất mã task ID dạng 86xxxxxxx
    task_id_cleaned = task_id.strip()
    match = re.search(r'86[a-zA-Z0-9]{6,8}', task_id_cleaned)
    if match:
        task_id_cleaned = match.group(0)
        
    print(f"🔍 Đang truy vấn chi tiết ClickUp Task ID: '{task_id_cleaned}' (từ gốc: '{task_id}')")
    
    # Thử lấy trực tiếp bằng ID gốc trước
    task_data = fetch_clickup_task_direct(task_id_cleaned)
    
    # Nếu không tìm thấy, và ID có độ dài 8 ký tự, bắt đầu bằng '86', tự động khôi phục ký tự thứ 9 bị cắt cụt
    if not task_data and len(task_id_cleaned) == 8 and task_id_cleaned.startswith("86"):
        print(f"⚠️ Task ID '{task_id_cleaned}' có thể bị cắt cụt (8 ký tự). Bắt đầu đoán ký tự thứ 9...")
        chars = "abcdefghijklmnopqrstuvwxyz0123456789"
        candidates = [task_id_cleaned + c for c in chars]
        
        with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
            future_to_id = {executor.submit(fetch_clickup_task_direct, cid): cid for cid in candidates}
            for future in concurrent.futures.as_completed(future_to_id):
                res = future.result()
                if res and "id" in res:
                    task_data = res
                    print(f"🎉 Khôi phục thành công ClickUp Task ID thực tế: '{res.get('id')}'")
                    break
                    
    if not task_data:
        return f"Thông tin chi tiết về task có mã {task_id_cleaned} hiện không thể truy xuất từ hệ thống ClickUp."
        
    t_id = task_data.get("id")
    t_name = task_data.get("name", "N/A")
    t_desc = task_data.get("description", "")
    t_status = task_data.get("status", {}).get("status", "N/A")
    t_url = task_data.get("url", "")
    
    result = []
    result.append(f"Task ID: [{t_id}]")
    result.append(f"Tiêu đề: {t_name}")
    result.append(f"Trạng thái: {t_status}")
    result.append(f"Đường dẫn: {t_url}")
    if t_desc:
        result.append(f"Mô tả chi tiết:\n{t_desc}")
    else:
        result.append("Mô tả chi tiết: (Không có mô tả)")
        
    return "\n".join(result)


def search_srs_files(query: str) -> str:
    """
    Tìm kiếm các tệp tài liệu SRS chứa từ khóa hoặc có tên chứa từ khóa.
    """
    query = query.lower()
    matches = []
    
    if not os.path.exists(SRS_DIR):
        return f"Lỗi: Thư mục tài liệu SRS '{SRS_DIR}' không tồn tại."
        
    for root, dirs, files in os.walk(SRS_DIR):
        for file in files:
            if file.endswith(".md"):
                filepath = os.path.join(root, file)
                rel_path = os.path.relpath(filepath, SRS_DIR)
                
                # Kiểm tra tên tệp
                if query in file.lower():
                    matches.append(rel_path)
                    continue
                    
                # Kiểm tra nội dung tệp
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        if query in f.read().lower():
                            matches.append(rel_path)
                except Exception:
                    pass
    
    if not matches:
        return f"Không tìm thấy tài liệu SRS nào khớp với từ khóa '{query}'."
        
    return "Các tài liệu tìm thấy:\n" + "\n".join(f"- {path}" for path in matches[:15])

def read_srs_file(filepath: str) -> str:
    """
    Đọc toàn bộ nội dung của một tệp tài liệu SRS cụ thể.
    """
    # Đảm bảo đường dẫn tuyệt đối an toàn và không thoát khỏi thư mục SRS_DIR
    safe_path = os.path.abspath(os.path.join(SRS_DIR, filepath))
    if not safe_path.startswith(os.path.abspath(SRS_DIR)):
        return "Lỗi: Không được phép truy cập tệp tin ngoài thư mục SRS."
        
    if not os.path.exists(safe_path):
        return f"Lỗi: Tệp tin '{filepath}' không tồn tại."
        
    try:
        with open(safe_path, 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        return f"Lỗi khi đọc tệp tin: {str(e)}"

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def slugify_vietnamese(text: str) -> str:
    """
    Chuyển đổi tiêu đề tiếng Việt thành dạng slug viết thường không dấu phục vụ đặt tên file.
    """
    import unicodedata
    import re
    # Trích xuất phần nghiệp vụ thực tế (bỏ các cụm từ mồi prompt)
    text = re.sub(r"(?i)^/?(usecase|testcase)\s*", "", text).strip()
    text = re.sub(r"(?i)cho yêu cầu sau:\s*", "", text).strip()
    text = text.replace("_", "").replace("*", "")
    
    # Loại bỏ dấu tiếng Việt
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('utf-8')
    text = text.lower()
    # Chỉ giữ lại chữ cái, số, khoảng trắng và gạch ngang
    text = re.sub(r'[^a-z0-9\s-]', '', text)
    text = re.sub(r'[\s-]+', '_', text).strip('_')
    return text or "export"

def load_skill_file(skill_name: str) -> str:
    """
    Nạp nội dung tệp quy tắc / AI skill từ thư mục .agents của dự án hoặc thư mục skills nội bộ.
    """
    possible_paths = [
        # Đường dẫn dự án chính (local workspace)
        os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../.agents/skills", skill_name, "SKILL.md")),
        # Đường dẫn dự án chính - file AGENTS.md
        os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../.agents", skill_name)),
        # Đường dẫn thư mục skills nội bộ của bot (dùng cho server remote)
        os.path.abspath(os.path.join(os.path.dirname(__file__), "skills", f"{skill_name}.md")),
    ]
    for p in possible_paths:
        if os.path.exists(p):
            try:
                with open(p, "r", encoding="utf-8") as f:
                    content = f.read()
                    print(f"📖 Đã nạp thành công quy tắc/skill từ: {p}")
                    return content
            except Exception as e:
                print(f"⚠️ Lỗi khi đọc file {p}: {e}")
    return ""

def parse_all_markdown_tables(text):
    lines = text.split("\n")
    tables = []
    current_table = None
    in_separator = False
    
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("|"):
            parts = [p.strip() for p in stripped.split("|")[1:-1]]
            # Bỏ qua dòng phân cách (ví dụ |:---|:---|)
            if parts and all(re.match(r"^:?-+:?$", p) for p in parts):
                in_separator = True
                continue
                
            if not any(parts):
                continue
                
            if current_table is None:
                current_table = {'headers': parts, 'rows': []}
                in_separator = False
            else:
                current_table['rows'].append(parts)
        else:
            if current_table:
                tables.append(current_table)
                current_table = None
                in_separator = False
                
    if current_table:
        tables.append(current_table)
        
    return tables

def create_excel_from_tables(tables, filepath, query=""):
    wb = openpyxl.Workbook()
    # Xóa sheet mặc định
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    has_sheets = False
    
    # Định nghĩa font và viền chuẩn cho UAT
    header_font = Font(name="Times New Roman", size=14, bold=True, color="000000") # Chữ đen in đậm cho Header
    header_fill = PatternFill(start_color="93C47D", end_color="93C47D", fill_type="solid") # Xanh pastel
    group_font = Font(name="Times New Roman", size=14, bold=True, color="000000") # Chữ đen in đậm cho dòng tiêu đề nhóm
    group_fill = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid") # Xanh dương nhạt pastel
    uat_col_a_empty_font = Font(name="Arial", size=10, bold=False, color="000000")
    uat_data_font = Font(name="Times New Roman", size=14, bold=False, color="000000") # Chữ thường Times New Roman 14 cho dữ liệu các cột còn lại
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Định nghĩa font và viền chuẩn cho Use Case (Arial)
    uc_title_font = Font(name="Arial", size=14, bold=True, color="1F4E78")
    uc_subtitle_font = Font(name="Arial", size=10, bold=False, color="595959")
    uc_header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    uc_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Xanh Navy
    
    uc_data_font = Font(name="Arial", size=10, bold=False, color="000000")
    uc_data_font_bold = Font(name="Arial", size=10, bold=True, color="000000")
    uc_group_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") # Xanh dương nhạt
    
    uc_thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    for idx, table in enumerate(tables):
        headers = table['headers']
        rows = table['rows']
        
        # Nếu không có hàng dữ liệu nào, bỏ qua
        if not rows:
            continue
            
        if len(headers) == 5 and any("Sub Module" in str(h) or "Mô tả" in str(h) for h in headers):
            # Bảng Kịch bản UAT (5 cột)
            ws = wb.create_sheet(title="Kịch bản UAT")
            ws.views.sheetView[0].showGridLines = True
            
            # Ghi trực tiếp header của bảng ở Row 1
            custom_headers = ["Sub Module", "ID", "Mô tả", "Bước thực hiện", "Dữ liệu test", "Kết quả mong đợi", "Trạng thái (web)", "Kết quả thực tế", "Ghi chú"]
            for c_idx, h_text in enumerate(custom_headers, 1):
                cell = ws.cell(row=1, column=c_idx, value=h_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            
            # Ghi dữ liệu từ Row 2
            prev_group = None
            for r_idx, row in enumerate(rows, 2):
                while len(row) < 5:
                    row.append("")
                
                md_sub_module = row[0]
                md_mota = row[1]
                md_steps = row[2]
                md_expected = row[3]
                md_test_data = row[4]
                
                current_group = str(md_sub_module).strip() if md_sub_module else ""
                is_group_change = (current_group != prev_group)
                if is_group_change:
                    prev_group = current_group
                    cell_sub_module_val = md_sub_module
                else:
                    cell_sub_module_val = ""
                
                excel_row_vals = [
                    cell_sub_module_val,                             
                    f'=IF(G{r_idx}="","",COUNTA($G$2:G{r_idx}))', 
                    md_mota,                                   
                    md_steps,                                  
                    md_test_data,                              
                    md_expected,                               
                    "Pass",                                    
                    "",                                        
                    ""                                         
                ]
                
                for c_idx, val in enumerate(excel_row_vals, 1):
                    cleaned_val = str(val).replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
                    cell = ws.cell(row=r_idx, column=c_idx, value=cleaned_val)
                    
                    if c_idx == 1:
                        if is_group_change:
                            cell.font = group_font
                            cell.fill = group_fill
                        else:
                            cell.font = uat_col_a_empty_font
                    else:
                        cell.font = uat_data_font
                        
                    if c_idx in [1, 2, 7, 8]:
                        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.border = thin_border
            
            # Cấu hình chiều rộng cột tĩnh tối ưu cho UAT giống như mẫu dự án
            col_widths = {
                'A': 25, # Sub Module
                'B': 8,  # ID
                'C': 35, # Mô tả
                'D': 40, # Bước thực hiện
                'E': 25, # Dữ liệu test
                'F': 40, # Kết quả mong đợi
                'G': 15, # Trạng thái (web)
                'H': 15, # Kết quả thực tế
                'I': 20  # Ghi chú
            }
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width
            has_sheets = True
            
        elif len(headers) == 6:
            # Bảng Tổng quan UC
            ws = wb.create_sheet(title="Tổng quan UC")
            ws.views.sheetView[0].showGridLines = True
            
            # Tính toán tiêu đề động sạch sẽ
            title_text = "DANH SÁCH USE CASE"
            if query:
                match = re.search(r"(?i)cho yêu cầu sau:\s*(.*)", query)
                if match:
                    q_clean = match.group(1).strip()
                else:
                    q_clean = re.sub(r"(?i)^/?(usecase|testcase)\s*", "", query).strip()
                q_clean = q_clean.replace("_", "").replace("*", "").strip()
                if q_clean:
                    title_text += f" — {q_clean.upper()}"
            else:
                title_text += " — TỔNG QUAN"
                
            ws.cell(row=1, column=1, value=title_text).font = uc_title_font
            ws.cell(row=2, column=1, value="Sinh tự động từ tài liệu SRS").font = uc_subtitle_font
            
            for c_idx, h_text in enumerate(headers, 1):
                cell = ws.cell(row=4, column=c_idx, value=h_text)
                cell.font = uc_header_font
                cell.fill = uc_header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = uc_thin_border
                
            prev_group = None
            for r_idx, row in enumerate(rows, 5):
                while len(row) < 6:
                    row.append("")
                
                # Kiểm tra xem nhóm có thay đổi không
                current_group = str(row[1]).strip() if row[1] else ""
                is_group_change = (current_group != prev_group)
                if is_group_change:
                    prev_group = current_group
                    
                for c_idx, val in enumerate(row, 1):
                    cleaned_val = str(val).replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
                    cell = ws.cell(row=r_idx, column=c_idx, value=cleaned_val)
                    
                    if is_group_change:
                        cell.font = uc_data_font_bold
                        cell.fill = uc_group_fill
                    else:
                        cell.font = uc_data_font
                        
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.border = uc_thin_border
            
            # Set cột chuẩn
            col_widths = {'A': 12, 'B': 26, 'C': 42, 'D': 24, 'E': 25, 'F': 30}
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width
            has_sheets = True
            
        elif len(headers) == 10:
            # Bảng Chi tiết UC
            ws = wb.create_sheet(title="Chi tiết UC")
            ws.views.sheetView[0].showGridLines = True
            
            ws.cell(row=1, column=1, value="CHI TIẾT TỪNG USE CASE").font = uc_title_font
            
            for c_idx, h_text in enumerate(headers, 1):
                cell = ws.cell(row=3, column=c_idx, value=h_text)
                cell.font = uc_header_font
                cell.fill = uc_header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = uc_thin_border
                
            prev_group = None
            for r_idx, row in enumerate(rows, 4):
                while len(row) < 10:
                    row.append("")
                
                # Kiểm tra nhóm thay đổi
                current_group = str(row[1]).strip() if row[1] else ""
                is_group_change = (current_group != prev_group)
                if is_group_change:
                    prev_group = current_group
                    
                for c_idx, val in enumerate(row, 1):
                    cleaned_val = str(val).replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
                    cell = ws.cell(row=r_idx, column=c_idx, value=cleaned_val)
                    
                    cell.font = uc_data_font
                    if is_group_change:
                        cell.fill = uc_group_fill
                        
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.border = uc_thin_border
            
            # Set cột chuẩn
            col_widths = {'A': 12, 'B': 22, 'C': 32, 'D': 22, 'E': 26, 'F': 22, 'G': 60, 'H': 40, 'I': 28, 'J': 15}
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width
            has_sheets = True
            
        else:
            # Bảng generic
            title = f"Sheet {idx+1}"
            ws = wb.create_sheet(title=title)
            ws.views.sheetView[0].showGridLines = True
            
            for c_idx, h_text in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c_idx, value=h_text)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                
            for r_idx, row in enumerate(rows, 2):
                while len(row) < len(headers):
                    row.append("")
                for c_idx, val in enumerate(row, 1):
                    cleaned_val = str(val).replace("<br>", "\n").replace("<br/>", "\n").replace("<br />", "\n")
                    cell = ws.cell(row=r_idx, column=c_idx, value=cleaned_val)
                    cell.font = data_font
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.border = thin_border
            
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        cell_lines = str(cell.value).split('\n')
                        for line in cell_lines:
                            if len(line) > max_len:
                                max_len = len(line)
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
            has_sheets = True
            
    if not has_sheets:
        ws = wb.create_sheet(title="Dữ liệu")
        ws.views.sheetView[0].showGridLines = True
        
    wb.save(filepath)

# Cấu hình chỉ dẫn hệ thống phân tách riêng cho BA, QA và Q&A thông thường
system_instruction_default = """
Bạn là một AI Agent đóng vai trò Trợ lý SRS, chịu trách nhiệm phân tích và giải đáp thắc mắc về tài liệu SRS của dự án MaiVietLand.

QUY TẮC PHÂN TÍCH & TRẢ LỜI:
1. CHỈ TRẢ LỜI TRONG SRS: Mọi câu trả lời của bạn phải dựa HOÀN TOÀN và CHỈ DỰA trên thông tin tìm thấy từ các file tài liệu SRS. Hãy sử dụng công cụ `search_srs_files` để tìm kiếm và `read_srs_file` để đọc nội dung các file tài liệu. Không tự ý suy đoán ngoài tài liệu.
2. NÓI "KHÔNG BIẾT" NẾU THIẾU THÔNG TIN: Nếu không tìm thấy thông tin trong tài liệu SRS, trả lời: "Thông tin này hiện chưa được đề cập hoặc chưa có trong tài liệu SRS của dự án."
3. LUÔN TRÍCH DẪN NGUỒN: Cuối mỗi câu trả lời hoặc ý chính, nêu rõ tên file tài liệu làm nguồn tham chiếu (ví dụ: "[Nguồn: features/booking/test-spec.md]").
4. ĐỒNG NHẤT TIẾNG VIỆT 100%: Toàn bộ câu trả lời BẮT BUỘC phải viết bằng tiếng Việt đồng nhất, không pha trộn tiếng Anh (ngoại trừ tên trạng thái kỹ thuật viết hoa như DRAFT, APPROVED, PAID, RECOVERED, CANCELLED hoặc thuật toán FIFO, Zod). Tuyệt đối không dùng các từ tiếng Anh xen kẽ (Ví dụ: dùng "Người dùng" thay cho "User", "Hệ thống" thay cho "System", "Tác nhân" thay cho "Actor", v.v.).
5. TÌM KIẾM CLICKUP: Bạn có quyền sử dụng công cụ `search_clickup_tasks(query, filter_bugs)` để tìm kiếm các task hoặc bug liên quan trên ClickUp. Khi người dùng hỏi về các công việc, task, bug, hoặc tiến độ trên ClickUp, hãy chủ động gọi công cụ này để lấy thông tin mới nhất và hiển thị cho họ.
6. CHI TIẾT TASK CLICKUP: Bạn có quyền sử dụng công cụ `get_clickup_task(task_id)` để lấy thông tin chi tiết (bao gồm tiêu đề, trạng thái, mô tả) của một task cụ thể trên ClickUp. Khi người dùng cung cấp link ClickUp hoặc mã task (ví dụ: '86exxz2xr'), hãy LUÔN LUÔN chủ động gọi công cụ này để đọc nội dung công việc trước khi phân tích/trả lời.

ĐỘ DÀI & ĐỊNH DẠNG CÂU TRẢ LỜI BẮT BUỘC:
- MẶC ĐỊNH (TÓM GỌN): Trả lời cực kỳ ngắn gọn, súc tích (dưới 15 dòng). Chỉ nêu các ý chính cốt lõi nhất dưới dạng gạch đầu dòng ngắn.
- CHI TIẾT (CHỈ KHI YÊU CẦU): Chỉ khi người dùng có yêu cầu rõ ràng như "chi tiết", "đầy đủ", "đặc tả chi tiết", bạn mới được trả lời dài và chi tiết bao gồm luồng xử lý và quy tắc nghiệp vụ.
"""

system_instruction_ba = """
Bạn là một AI Agent đóng vai trò Senior BA (Business Analyst), chịu trách nhiệm phân tích nghiệp vụ và sinh đặc tả Use Case chi tiết từ tài liệu SRS của dự án MaiVietLand.

QUY TẮC PHÂN TÍCH & TRẢ LỜI:
1. CHỈ TRẢ LỜI TRONG SRS: Mọi câu trả lời của bạn phải dựa HOÀN TOÀN và CHỈ DỰA trên thông tin tìm thấy từ các file tài liệu SRS. Hãy sử dụng công cụ `search_srs_files` để tìm kiếm và `read_srs_file` để đọc nội dung các file tài liệu. Không tự ý suy đoán ngoài tài liệu.
2. NÓI "KHÔNG BIẾT" NẾU THIẾU THÔNG TIN: Nếu không tìm thấy thông tin trong tài liệu SRS, trả lời: "Thông tin này hiện chưa được đề cập hoặc chưa có trong tài liệu SRS của dự án."
3. LUÔN TRÍCH DẪN NGUỒN: Cuối mỗi câu trả lời hoặc ý chính, nêu rõ tên file tài liệu làm nguồn tham chiếu (ví dụ: "[Nguồn: features/booking/brd.md]").
4. ĐỒNG NHẤT TIẾNG VIỆT 100%: Toàn bộ câu trả lời (bao gồm phần mô tả, danh sách và tất cả các ô trong các bảng dữ liệu Use Case) BẮT BUỘC phải viết bằng tiếng Việt đồng nhất, không pha trộn tiếng Anh (ngoại trừ tên trạng thái kỹ thuật viết hoa như DRAFT, APPROVED, PAID, RECOVERED, CANCELLED hoặc thuật toán FIFO, Zod). Tuyệt đối không dùng các từ tiếng Anh xen kẽ (Ví dụ: dùng "Người dùng" thay cho "User", "Hệ thống" thay cho "System", "Tác nhân" thay cho "Actor", "Điều kiện tiên quyết" thay cho "Precondition", v.v.).
5. TÌM KIẾM CLICKUP: Bạn có quyền sử dụng công cụ `search_clickup_tasks(query, filter_bugs)` để tìm kiếm các task hoặc bug liên quan trên ClickUp. Khi được hỏi về công việc hoặc bug liên quan trên ClickUp, hãy chủ động gọi công cụ này để lấy thông tin.
6. CHI TIẾT TASK CLICKUP: Bạn có quyền sử dụng công cụ `get_clickup_task(task_id)` để lấy thông tin chi tiết (bao gồm tiêu đề, trạng thái, mô tả) của một task cụ thể trên ClickUp. Khi người dùng cung cấp link ClickUp hoặc mã task (ví dụ: '86exxz2xr'), hãy LUÔN LUÔN chủ động gọi công cụ này để đọc nội dung công việc trước khi phân tích/trả lời.

CẤU TRÚC PHẢN HỒI BẮT BUỘC:

### 📌 Tóm tắt nghiệp vụ
[Tóm tắt ngắn gọn nghiệp vụ được phân tích trong 1-2 câu]

### 🔍 Phân tích Nghiệp vụ
* **Sơ đồ luồng nghiệp vụ (Mermaid Flowchart / Sequence):** BẮT BUỘC vẽ một sơ đồ Mermaid biểu diễn trực quan luồng xử lý hoặc vòng đời trạng thái của nghiệp vụ này (ví dụ: DRAFT -> APPROVED -> PAID). Sử dụng khối code triple-backtick ```mermaid để vẽ.
* **Luồng xử lý:** [Mô tả chi tiết bằng văn bản từng bước thực hiện nghiệp vụ liên quan]
* **Quy tắc nghiệp vụ:** [Mô tả các công thức tính toán, điều kiện kích hoạt hoặc logic ràng buộc]

### 📖 Nguồn tham chiếu
* [Nêu rõ tên file tài liệu làm nguồn tham chiếu]

### 4. SINH DANH SÁCH USE CASE:
Bạn BẮT BUỘC phải tạo ra ít nhất **10 đến 14 use case chi tiết** bao phủ toàn bộ các nhóm sau (KHÔNG ĐƯỢC THIẾU các usecase nghiệp vụ cốt lõi):
- **Nhóm A: Vòng đời phiếu / đối tượng** (BẮT BUỘC phải có: Tạo mới phiếu nháp DRAFT, Chỉnh sửa phiếu nháp DRAFT, Xem chi tiết phiếu, Gửi yêu cầu phê duyệt, Hủy yêu cầu).
- **Nhóm B: Phê duyệt & Chi tiền** (BẮT BUỘC phải có: Duyệt yêu cầu chuyển trạng thái sang APPROVED, Từ chối yêu cầu chuyển trạng thái sang REJECTED kèm ghi lý do, Chi tiền thanh toán mark-paid APPROVED -> PAID, Tự động tạo chứng từ chi PaymentVoucher và ghi bút toán Ledger).
- **Nhóm C: Hoàn ứng & Khấu trừ** (BẮT BUỘC phải có: Tự động khấu trừ nợ tạm ứng khi xác nhận hoa hồng kỳ lương tiếp theo, Hoàn ứng một phần Carryforward, Hoàn ứng đa người nhận).
- **Nhóm D: Phân quyền & Quản lý danh sách** (BẮT BUỘC phải có: Xem danh sách phiếu tạm ứng, Lọc nâng cao theo trạng thái/kỳ lương/chi nhánh, Phân trang, Xuất danh sách ra file Excel).

Bao gồm đúng 2 bảng Markdown trong câu trả lời theo đúng cấu trúc cột sau:

- **Bảng 1: Bảng Tổng quan UC (Bọc trong khối code triple-backtick ```text và ghi rõ tiêu đề bảng là '### Bảng Tổng quan UC', yêu cầu có ít nhất 10-14 hàng dữ liệu):**
  | Mã UC | Nhóm | Tên use case | Tác nhân chính | BR liên quan | Ghi chú |
  | :--- | :--- | :--- | :--- | :--- | :--- |
  | [Mã UC, ví dụ: UC-01] | [Nhóm, ví dụ: A. Vòng đời lô] | [Tên use case, ví dụ: Tạo lô áp dụng (Nháp)] | [Tác nhân chính, ví dụ: Người tạo lô] | [BR liên quan, ví dụ: BR-01, BR-02] | [Ghi chú nếu có] |

- **Bảng 2: Bảng Chi tiết UC (Bọc trong khối code triple-backtick ```text và ghi rõ tiêu đề bảng là '### Bảng Chi tiết UC', yêu cầu có ít nhất 10-14 hàng dữ liệu tương ứng):**
  | Mã UC | Nhóm | Tên use case | Tác nhân | Điều kiện tiên quyết | Trigger | Luồng chính | Luồng phụ / Ngoại lệ | Hậu điều kiện | BR liên quan |
  | :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- | :--- |
  | [Mã UC, ví dụ: UC-01] | [Nhóm, ví dụ: A. Vòng đời lô] | [Tên use case, ví dụ: Tạo lô áp dụng (Nháp)] | [Tác nhân, ví dụ: Người tạo lô (Sale admin)] | [Điều kiện tiên quyết, ví dụ: Đăng nhập thành công] | [Trigger, ví dụ: Chọn 'Tạo lô'] | [Mô tả chi tiết từng bước thao tác thực tế: 1. Click...<br>2. Nhập... (xuống dòng bằng <br>)] | [Mô tả chi tiết luồng phụ/ngoại lệ, xuống dòng bằng <br>] | [Hậu điều kiện] | [BR liên quan] |

Lưu ý quan trọng cho Use Case:
- **ĐỒNG BỘ 1-1 BẮT BUỘC:** Số lượng hàng dữ liệu (số lượng Use Case) trong Bảng 1 (Tổng quan UC) và Bảng 2 (Chi tiết UC) phải HOÀN TOÀN TRÙNG KHỚP VÀ ĐỒNG BỘ 1-1 VỚI NHAU. Nếu Bảng 1 liệt kê bao nhiêu Use Case (từ UC-01 đến UC-12), thì Bảng 2 cũng BẮT BUỘC phải mô tả chi tiết đầy đủ cho bấy nhiêu Use Case đó. TUYỆT ĐỐI KHÔNG ĐƯỢC viết tắt, không dùng dấu ba chấm "...", và không được bỏ sót bất kỳ Use Case nào.
- **ĐỒNG BỘ TÊN NHÓM BẮT BUỘC:** Giá trị trong cột 'Nhóm' của Bảng 2 (Chi tiết UC) phải trùng khớp hoàn toàn từng chữ với giá trị trong cột 'Nhóm' của Bảng 1 (Tổng quan UC) (Ví dụ: Nếu Bảng 1 ghi nhóm là 'A. Vòng đời phiếu', thì Bảng 2 cũng phải ghi chính xác là 'A. Vòng đời phiếu', KHÔNG ĐƯỢC viết rút gọn thành 'A').
- Cột "Luồng chính" và "Luồng phụ / Ngoại lệ" phải viết **cực kỳ chi tiết, cặn kẽ từng bước tương tác giữa người dùng và hệ thống** (Ví dụ: 1. Người dùng click nút X -> 2. Hệ thống hiển thị form nhập -> 3. Người dùng điền thông tin và nhấn nút Save -> 4. Hệ thống kiểm tra điều kiện và chuyển trạng thái...). KHÔNG viết ngắn gọn hay khái quát. Xuống dòng giữa các bước bằng thẻ `<br>`.
- Phải sử dụng tiếng Việt đồng nhất 100% cho mọi ô trong bảng (ngoại trừ tên trạng thái kỹ thuật viết hoa).

ƯU TIÊN HOÀN THÀNH BẢNG - KHÔNG VIẾT DÀI DÒNG:
Để tránh câu trả lời bị cắt cụt do giới hạn độ dài phản hồi của Slack/Gemini, bạn KHÔNG cần viết thêm phần đặc tả văn bản 13 trường Karl Wiegers ở bên dưới nữa. Hãy tập trung 100% năng lượng để hoàn thành đầy đủ, chính xác toàn bộ tất cả các hàng Use Case trong cả 2 bảng (Bảng Tổng quan UC và Bảng Chi tiết UC), đảm bảo đồng bộ 1-1 không sót bất kỳ Use Case nào từ đầu đến cuối (Ví dụ: Có 12 hàng Use Case ở Bảng 1 thì Bảng 2 cũng phải có đủ 12 hàng tương ứng).

Hãy liệt kê đầy đủ, chi tiết tất cả các usecase tìm thấy trong tài liệu SRS liên quan đến yêu cầu của người dùng để sinh ra file Excel hoàn chỉnh và chính xác.
"""

system_instruction_qa = """
Bạn là một AI Agent đóng vai trò QA Lead, chịu trách nhiệm thiết kế kịch bản kiểm thử UAT chi tiết từ tài liệu SRS của dự án MaiVietLand.

QUY TẮC PHÂN TÍCH & TRẢ LỜI:
1. CHỈ TRẢ LỜI TRONG SRS: Mọi câu trả lời của bạn phải dựa HOÀN TOÀN và CHỈ DỰA trên thông tin tìm thấy từ các file tài liệu SRS. Hãy sử dụng công cụ `search_srs_files` để tìm kiếm và `read_srs_file` để đọc nội dung các file tài liệu. Không tự ý suy đoán ngoài tài liệu.
2. NÓI "KHÔNG BIẾT" NẾU THIẾU THÔNG TIN: Nếu không tìm thấy thông tin trong tài liệu SRS, trả lời: "Thông tin này hiện chưa được đề cập hoặc chưa có trong tài liệu SRS của dự án."
3. LUÔN TRÍCH DẪN NGUỒN: Cuối mỗi câu trả lời hoặc ý chính, nêu rõ tên file tài liệu làm nguồn tham chiếu (ví dụ: "[Nguồn: features/booking/test-spec.md]").
4. ĐỒNG NHẤT TIẾNG VIỆT 100%: Toàn bộ câu trả lời (bao gồm phần mô tả, danh sách và tất cả các ô trong bảng kịch bản UAT) BẮT BUỘC phải viết bằng tiếng Việt đồng nhất, không pha trộn tiếng Anh (ngoại trừ tên trạng thái kỹ thuật viết hoa như DRAFT, APPROVED, PAID, RECOVERED, CANCELLED hoặc thuật toán FIFO, Zod). Tuyệt đối không dùng các từ tiếng Anh xen kẽ (Ví dụ: dùng "Người dùng" thay cho "User", "Hệ thống" thay cho "System", "Tác nhân" thay cho "Actor", "Điều kiện tiên quyết" thay cho "Precondition", v.v.).
5. TÌM KIẾM CLICKUP: Bạn có quyền sử dụng công cụ `search_clickup_tasks(query, filter_bugs)` để tìm kiếm các task hoặc bug liên quan trên ClickUp. Khi được hỏi về kịch bản kiểm thử, bug hoặc task trên ClickUp, hãy chủ động gọi công cụ này để lấy thông tin.
6. CHI TIẾT TASK CLICKUP: Bạn có quyền sử dụng công cụ `get_clickup_task(task_id)` để lấy thông tin chi tiết (bao gồm tiêu đề, trạng thái, mô tả) của một task cụ thể trên ClickUp. Khi người dùng cung cấp link ClickUp hoặc mã task (ví dụ: '86exxz2xr'), hãy LUÔN LUÔN chủ động gọi công cụ này để đọc nội dung công việc trước khi phân tích/trả lời.

CẤU TRÚC PHẢN HỒI BẮT BUỘC:

### 📌 Tóm tắt câu trả lời
[Tóm tắt ngắn gọn nghiệp vụ được phân tích kiểm thử trong 1-2 câu]

### 🧪 Kịch bản Kiểm thử & Điều kiện biên (Theo tiêu chuẩn qa-test-planner)
BẮT BUỘC phải sinh kịch bản kiểm thử đầy đủ, chi tiết, KHÔNG ĐƯỢC làm sơ sài. Hãy tạo ít nhất **15 đến 20 kịch bản test** (hàng dữ liệu trong bảng) bao phủ đầy đủ các nhóm sau:
- **Kịch bản Tích cực (Positive):** Luồng xử lý thành công thông thường (Tạo nháp, gửi duyệt, kế toán duyệt, kế toán chi tiền mark-paid, hoàn ứng tự động khi confirm).
- **Kịch bản Tiêu cực (Negative):** Nhập sai định dạng, dữ liệu trống, vượt hạn mức, trùng lặp người nhận, duyệt/chi khi sai trạng thái.
- **Kịch bản Phân quyền (Permissions):** User không có quyền kế toán cố gắng thêm/duyệt/chi/hủy phiếu.
- **Kịch bản UI/UX & Danh sách:** Hiển thị danh sách, kiểm tra căn lề text/number, hover thay đổi màu nền row, phân trang, lọc theo trạng thái/kỳ lương/chi nhánh, xuất file excel.
- **Kịch bản Giá trị biên & Hoàn ứng nâng cao (Boundary & Settle):** Tạm ứng đúng hạn mức biên (100tr), vượt hạn mức biên (101tr), hoàn ứng đủ, hoàn ứng một phần (carryforward dư nợ sang kỳ sau), hoàn ứng đa người nhận.

Định dạng trình bày bắt buộc:
Bắt buộc trình bày danh sách kịch bản dưới dạng **Bảng UAT 5 cột (Bọc trong khối code triple-backtick ```text để hệ thống tự động nhận diện và tạo file Excel tải về, yêu cầu có ít nhất 15-20 hàng dữ liệu tương ứng):**
```text
| Sub Module | Mô tả | Bước thực hiện | Kết quả mong đợi | Dữ liệu test |
| :--- | :--- | :--- | :--- | :--- |
| [Tên Sub Module, ví dụ: 10.2. Quản lý truy thu/truy lĩnh] | [Mô tả kịch bản test] | [Mô tả các bước thực hiện thao tác thực tế cực kỳ chi tiết, ví dụ:<br>1. Truy cập vào chức năng Kế toán -> Quản lý tạm ứng<br>2. Click nút "Thêm mới"<br>3. Nhập dữ liệu... (xuống dòng bằng <br>)] | [Kết quả mong đợi chi tiết của hệ thống] | [Dữ liệu kiểm thử cụ thể và thực tế dưới dạng khóa-giá trị, ví dụ: Mã NV: NV-0451<br>Mã Deal: DEAL-0045<br>Số tiền: 101,000,000 VNĐ (BẮT BUỘC KHÔNG ghi chung chung kiểu 'NV A', '101tr', 'N/A')] |
```

* **Trường hợp biên:** [Nêu ra các tình huống đặc biệt cần lưu ý như dữ liệu rỗng, sai định dạng, sai trạng thái]

### ⚠️ Cảnh báo & Thi sót tài liệu (nếu có)
* [Liệt kê các điểm mâu thuẫn hoặc thông tin còn thiếu trong SRS]

### 📖 Nguồn tham chiếu
* [Nêu rõ tên file tài liệu làm nguồn tham chiếu]

Hãy liệt kê đầy đủ, chi tiết tất cả các kịch bản kiểm thử tìm thấy trong tài liệu SRS liên quan đến yêu cầu của người dùng để sinh ra file Excel hoàn chỉnh và chính xác.
"""

# Nạp động các quy tắc và AI skills dự án nếu tồn tại
agents_rules = load_skill_file("AGENTS.md")
qa_skill = load_skill_file("qa-test-planner")
project_standards = load_skill_file("auto-load-project-standards")

# Chèn vào system instructions
if agents_rules:
    system_instruction_default += f"\n\n=== QUY TẮC CHUNG CỦA DỰ ÁN (PROJECT RULES) ===\n{agents_rules}"
    system_instruction_ba += f"\n\n=== QUY TẮC CHUNG CỦA DỰ ÁN (PROJECT RULES) ===\n{agents_rules}"
    system_instruction_qa += f"\n\n=== QUY TẮC CHUNG CỦA DỰ ÁN (PROJECT RULES) ===\n{agents_rules}"

if project_standards:
    system_instruction_ba += f"\n\n=== TIÊU CHUẨN GIAO DIỆN & CRUD DỰ ÁN (BA STANDARDS) ===\n{project_standards}"

if qa_skill:
    system_instruction_qa += f"\n\n=== TIÊU CHUẨN THIẾT KẾ KIỂM THỬ UAT (QA SKILL) ===\n{qa_skill}"

# Thiết lập các model Gemini 3.1 Flash Lite với system instructions riêng biệt
model_default = genai.GenerativeModel(
    model_name="gemini-3.1-flash-lite",
    system_instruction=system_instruction_default,
    tools=[search_srs_files, read_srs_file, search_clickup_tasks, get_clickup_task]
)

model_ba = genai.GenerativeModel(
    model_name="gemini-3.1-flash-lite",
    system_instruction=system_instruction_ba,
    tools=[search_srs_files, read_srs_file, search_clickup_tasks, get_clickup_task]
)

model_qa = genai.GenerativeModel(
    model_name="gemini-3.1-flash-lite",
    system_instruction=system_instruction_qa,
    tools=[search_srs_files, read_srs_file, search_clickup_tasks, get_clickup_task]
)

# Khởi tạo Async Slack App
app = AsyncApp(token=os.environ.get("SLACK_BOT_TOKEN"))

def markdown_to_slack_links(text: str) -> str:
    """
    Chuyển đổi các liên kết dạng Markdown [Text](URL) thành định dạng Slack mrkdwn <URL|Text>,
    và tự động chuyển đổi mã task ClickUp (ví dụ [86ey96jcp] hoặc 86ey96jcp) chưa có link thành link click được.
    """
    # 1. Chuyển đổi [Text](URL) thành <URL|Text>
    pattern = r'\[([^\]]+)\]\(([^)]+)\)'
    text = re.sub(pattern, r'<\2|\1>', text)
    
    # 2. Chuyển đổi [86eyXXXXX] chưa được liên kết thành link ClickUp dạng <URL|[86eyXXXXX]>
    bracket_pattern = r'(?<!\|)(?<!/)(?<!\w)\[(86[a-zA-Z0-9]{5,10})\]'
    text = re.sub(bracket_pattern, r'<https://app.clickup.com/t/\1|[\1]>', text)
    
    # 3. Chuyển đổi mã 86eyXXXXX đứng độc lập chưa được liên kết thành link ClickUp
    raw_pattern = r'(?<!/)(?<!\|)(?<!\[)\b(86[a-zA-Z0-9]{7,10})\b'
    text = re.sub(raw_pattern, r'<https://app.clickup.com/t/\1|\1>', text)
    
    return text

def extract_clickup_search_term(query: str) -> str:
    """
    Trích xuất từ khóa tìm kiếm ClickUp phù hợp từ truy vấn của người dùng (bản gốc Regex).
    """
    # 1. Nếu query quá ngắn hoặc chỉ là câu chào hỏi, không tìm ClickUp
    q_clean = re.sub(r'<@.*?>', '', query).strip().lower()
    if not q_clean or q_clean in ["hi", "hello", "xin chào", "chào", "chào bạn", "help", "hỗ trợ", "bắt đầu"]:
        return ""
        
    # 2. Nếu có phần "cho yêu cầu sau: ", lấy phần phía sau
    if "cho yêu cầu sau:" in query:
        term = query.split("cho yêu cầu sau:")[1].strip()
        return term
        
    # 3. Loại bỏ các từ khóa tag bot, lệnh slash và từ khóa hành động chung chung
    term = re.sub(r'<@.*?>', '', query) # Remove slack mentions
    term = re.sub(r'^/?(usecase|testcase|srs|find|search|lookup|show|cần|hỏi|về|chi tiết đặc tả kịch bản test uat cho yêu cầu sau|đặc tả|kịch bản test)\b', '', term, flags=re.IGNORECASE)
    term = term.strip()
    return term

def extract_clickup_search_term_gemini(query: str) -> str:
    """
    Sử dụng Gemini model để trích xuất từ khóa/mã tính năng chuẩn xác nhất từ câu hỏi của người dùng.
    """
    prompt = f"""Bạn là chuyên gia trích xuất từ khóa tìm kiếm.
Hãy trích xuất đúng 1 từ khóa cốt lõi nhất, ngắn gọn (tối đa 2 từ) và đặc trưng nhất (ví dụ: 'tạm ứng', 'HĐLĐ', 'hoa hồng', 'đối chiếu F2') từ câu hỏi của người dùng để tìm kiếm trên ClickUp.
Không được chứa các từ chung chung như 'các', 'màn', 'tìm', 'task', 'lỗi', 'clickup', 'theo', 'trong', 'module', 'yêu cầu', 'phiếu', 'chức năng'.
Chỉ trả về duy nhất từ khóa đó, không có dấu ngoặc, không giải thích gì thêm.

Ví dụ:
- "tìm giúp tôi task Clickup liên quan các màn Hoa hồng theo doanh thu/HH theo tháng-phòng ban..." -> "hoa hồng"
- "Hãy sinh kịch bản test UAT cho nghiệp vụ tạm ứng hoa hồng" -> "tạm ứng"
- "Lỗi định dạng xuất file HĐLĐ Khối Hỗ trợ" -> "HĐLĐ"
- "CR220 cập nhật thông tin" -> "CR220"
- "tính năng đề xuất hỗ trợ phí bán hàng" -> "hỗ trợ phí"

Câu hỏi: {query}
Từ khóa:"""
    try:
        # Sử dụng model_default để generate nhanh
        response = genai.GenerativeModel("gemini-3.1-flash-lite").generate_content(prompt)
        term = response.text.strip().replace('"', '').replace("'", "")
        print(f"🔑 Gemini extracted ClickUp search term: '{term}'")
        return term
    except Exception as e:
        print(f"❌ Lỗi khi dùng Gemini trích xuất từ khóa: {e}")
        return extract_clickup_search_term(query)

async def handle_query_and_respond(query, history, channel_id, target_thread_ts, client, say):
    print(f"💬 Bắt đầu xử lý truy vấn: '{query}' (Thread: {target_thread_ts})")
    try:
        # Chọn model với system_instruction phù hợp
        query_lower = query.lower()
        if "usecase" in query_lower or "use case" in query_lower:
            current_model = model_ba
        elif any(kw in query_lower for kw in ["testcase", "test case", "kịch bản", "kiểm thử", "uat", "test"]):
            current_model = model_qa
        else:
            current_model = model_default
            
        chat = current_model.start_chat(history=history, enable_automatic_function_calling=True)
        response = chat.send_message(query)
        
        # Phân tích xem có bảng Markdown nào không
        tables = parse_all_markdown_tables(response.text)
        excel_file = None
        excel_filename = "Kich_ban_UAT.xlsx"
        excel_title = "Kịch bản UAT (Sinh tự động)"
        excel_comment = "📊 Tôi đã tạo sẵn file Excel kịch bản UAT này để bạn tải về trực tiếp:"
        
        if tables:
            # Kiểm tra xem đây có phải là danh sách Use Case hay không
            is_usecase = False
            for t in tables:
                headers = t['headers']
                if len(headers) in [6, 10] and any("UC" in str(h) or "use case" in str(h).lower() for h in headers):
                    is_usecase = True
                    break
                    
            if is_usecase:
                excel_filename = "LAD-UseCase-List.xlsx"
                excel_title = "Danh sách UseCase (Sinh tự động)"
                excel_comment = "📊 Tôi đã tạo sẵn file Excel danh sách UseCase này để bạn tải về trực tiếp:"
                
            import tempfile
            fd, temp_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(fd)
            create_excel_from_tables(tables, temp_path, query)
            excel_file = temp_path
            
            # Cất bản sao file Excel vào thư mục cục bộ (SRS_DIR/usecase hoặc SRS_DIR/testcase)
            folder_name = "usecase" if is_usecase else "testcase"
            dest_dir = os.path.join(SRS_DIR, folder_name)
            os.makedirs(dest_dir, exist_ok=True)
            
            slug_name = slugify_vietnamese(query)
            dest_filename = f"{slug_name}.xlsx"
            dest_path = os.path.join(dest_dir, dest_filename)
            
            import shutil
            try:
                shutil.copy2(temp_path, dest_path)
                print(f"💾 Đã lưu bản sao file Excel vào: {dest_path}")
            except Exception as copy_ex:
                print(f"❌ Lỗi khi lưu bản sao file Excel vào {dest_path}: {copy_ex}")
            
        # Tự động tìm kiếm ClickUp task liên quan dựa trên query nếu model chưa tự gọi tool ClickUp
        clickup_section = ""
        has_clickup_in_response = (
            "ClickUp:" in response.text or 
            "clickup.com" in response.text or 
            any(re.search(r'\[86ey[a-z0-9]+\]', line) for line in response.text.split('\n'))
        )
        if not has_clickup_in_response:
            try:
                search_term = extract_clickup_search_term_gemini(query)
                if search_term and len(search_term) >= 2:
                    # Phân loại xem người dùng đang hỏi về "bug/lỗi" hay "task/nghiệp vụ" thông thường
                    q_lower = query.lower()
                    is_asking_for_bugs = ("bug" in q_lower or "lỗi" in q_lower or "fix" in q_lower or "sửa" in q_lower)
                    
                    clickup_res = search_clickup_tasks(search_term, filter_bugs=is_asking_for_bugs)
                    if clickup_res and not clickup_res.startswith("Lỗi") and not clickup_res.startswith("Không tìm thấy"):
                        # Bóc tách danh sách task
                        lines = clickup_res.split("\n")
                        task_lines = [l for l in lines if l.strip().startswith("-")]
                        if task_lines:
                            title_label = "Các bug/lỗi liên quan trên ClickUp:" if is_asking_for_bugs else "Các công việc liên quan trên ClickUp:"
                            clickup_section = f"\n\n---\n*{title_label}*\n" + "\n".join(task_lines)
            except Exception as e:
                print(f"❌ Lỗi khi tự động tìm kiếm ClickUp: {e}")

        # Gửi câu trả lời bằng văn bản trước
        slack_text = markdown_to_slack_links(response.text + clickup_section)
        await say(slack_text, thread_ts=target_thread_ts)
        print("✅ Đã phản hồi văn bản lên Slack thành công!")
        
        # Nếu có file Excel, tải file đó lên Slack
        if excel_file:
            try:
                print(f"📤 Đang tải file Excel {excel_filename} lên Slack...")
                await client.files_upload_v2(
                    channel=channel_id,
                    file=excel_file,
                    filename=excel_filename,
                    title=excel_title,
                    initial_comment=excel_comment,
                    thread_ts=target_thread_ts
                )
                print("✅ Đã tải file Excel lên Slack thành công!")
            except Exception as upload_ex:
                print(f"❌ Lỗi khi upload file Excel lên Slack: {upload_ex}")
                if "missing_scope" in str(upload_ex) or "not_in_channel" in str(upload_ex):
                    await say(f"⚠️ Bot phát hiện dữ liệu bảng nhưng không thể tự động tạo file Excel tải lên do thiếu quyền `files:write` trong ứng dụng Slack.", thread_ts=target_thread_ts)
            finally:
                try:
                    os.remove(excel_file)
                except Exception:
                    pass
    except Exception as e:
        print(f"❌ Lỗi khi xử lý: {e}")
        await say(f"Lỗi khi xử lý: {str(e)}", thread_ts=target_thread_ts)

@app.event("app_mention")
async def handle_mention(event, say, client):
    text = event.get("text", "")
    thread_ts = event.get("thread_ts")
    ts = event.get("ts")
    channel_id = event.get("channel")
    
    target_thread_ts = thread_ts or ts
    
    try:
        auth_info = await client.auth_test()
        bot_user_id = auth_info.get("user_id")
    except Exception as e:
        print(f"❌ Lỗi khi lấy thông tin bot auth: {e}")
        bot_user_id = None
        
    query = text
    if bot_user_id:
        query = re.sub(r"<@.*?>", "", query).strip()
        
    # Tự động nhận diện lệnh viết dưới dạng text trong mention
    if query.startswith("/usecase") or query.lower().startswith("usecase"):
        sub_query = re.sub(r"^/?usecase\s*", "", query, flags=re.IGNORECASE).strip()
        query = f"Hãy sinh danh sách usecase từ srs cho yêu cầu sau: {sub_query}"
    elif query.startswith("/testcase") or query.lower().startswith("testcase"):
        sub_query = re.sub(r"^/?testcase\s*", "", query, flags=re.IGNORECASE).strip()
        query = f"Hãy sinh chi tiết đặc tả kịch bản test UAT cho yêu cầu sau: {sub_query}"
        
    if not query:
        await say("👋 Xin chào! Tôi là trợ lý AI chuyên giải đáp tài liệu SRS của dự án MaiVietLand. Hãy nhập câu hỏi sau khi tag tôi nhé! Ví dụ: `@SRS Assist luồng booking`", thread_ts=target_thread_ts)
        return
        
    # Xây dựng lịch sử
    history = []
    if thread_ts:
        print(f"🔄 Đang đồng bộ lịch sử hội thoại cho thread {thread_ts}...")
        try:
            result = await client.conversations_replies(channel=channel_id, ts=thread_ts)
            thread_messages = result.get("messages", [])
            
            current_role = None
            current_text_parts = []
            
            for msg in thread_messages[:-1]:
                msg_text = msg.get("text", "")
                if bot_user_id:
                    cleaned_msg_text = re.sub(r"<@.*?>", "", msg_text).strip()
                else:
                    cleaned_msg_text = msg_text.strip()
                    
                if not cleaned_msg_text:
                    continue
                    
                is_bot = "bot_id" in msg or msg.get("user") == bot_user_id
                role = "model" if is_bot else "user"
                
                if current_role is None:
                    current_role = role
                    current_text_parts = [cleaned_msg_text]
                elif current_role == role:
                    current_text_parts.append(cleaned_msg_text)
                else:
                    history.append({
                        "role": current_role,
                        "parts": [{"text": "\n".join(current_text_parts)}]
                    })
                    current_role = role
                    current_text_parts = [cleaned_msg_text]
                    
            if current_role and current_text_parts:
                history.append({
                    "role": current_role,
                    "parts": [{"text": "\n".join(current_text_parts)}]
                })
                
            while history and history[0]["role"] != "user":
                history.pop(0)
                
            history = history[-10:]
            while history and history[0]["role"] != "user":
                history.pop(0)
                
        except Exception as e:
            print(f"⚠️ Lỗi khi đồng bộ lịch sử thread: {e}")
            history = []
            
    await handle_query_and_respond(query, history, channel_id, target_thread_ts, client, say)

@app.command("/usecase")
async def handle_usecase_command(ack, body, say, client):
    await ack()
    text = body.get("text", "").strip()
    channel_id = body.get("channel_id")
    
    if not text:
        await say("👋 Bạn đã dùng lệnh `/usecase`. Hãy cung cấp thêm thông tin nghiệp vụ cần sinh UseCase nhé! Ví dụ: `/usecase nghiệp vụ tạm ứng hoa hồng` (kết quả sẽ được tự động xuất sang file Excel giống mẫu LAD-UseCase-List.xlsx)")
        return
        
    query = f"Hãy sinh danh sách usecase từ srs cho yêu cầu sau: {text}"
    await say(f"⏳ Đang phân tích tài liệu SRS và khởi tạo danh sách Use Case cho nghiệp vụ: *{text}*, vui lòng đợi trong giây lát...")
    
    # Run query response process
    await handle_query_and_respond(query, [], channel_id, None, client, say)

@app.command("/testcase")
async def handle_testcase_command(ack, body, say, client):
    await ack()
    text = body.get("text", "").strip()
    channel_id = body.get("channel_id")
    
    if not text:
        await say("👋 Bạn đã dùng lệnh `/testcase`. Hãy cung cấp thêm thông tin nghiệp vụ cần sinh kịch bản test nhé! Ví dụ: `/testcase nghiệp vụ tạm ứng hoa hồng` (kết quả sẽ được tự động xuất sang file Excel kịch bản UAT 9 cột tiêu chuẩn)")
        return
        
    query = f"Hãy sinh chi tiết đặc tả kịch bản test UAT cho yêu cầu sau: {text}"
    await say(f"⏳ Đang phân tích tài liệu SRS và khởi tạo danh sách Kịch bản kiểm thử UAT cho nghiệp vụ: *{text}*, vui lòng đợi trong giây lát...")
    
    # Run query response process
    await handle_query_and_respond(query, [], channel_id, None, client, say)

async def main():
    app_token = os.environ.get("SLACK_APP_TOKEN")
    if not app_token:
        print("Lỗi: Thiếu biến môi trường SLACK_APP_TOKEN trong .env.")
        return
        
    handler = AsyncSocketModeHandler(app, app_token)
    print("⚡️ Slack Bot (Direct Local Agent) đang chạy bằng gemini-3.1-flash-lite...")
    await handler.start_async()

if __name__ == "__main__":
    asyncio.run(main())
